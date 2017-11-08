import openpyxl
from openpyxl import Workbook
import time

lista_grupo_12 = openpyxl.load_workbook('Lista Grupo 12.xlsx')
lista = (lista_grupo_12.get_sheet_by_name('Hoja1'))

lista_nombres=[]
lista_cedulas=[]
lista_apellidos=[]



                
nombres=lista["a2":"31"]
for filas in nombres:
        for nombre_wb in filas:
                lista_nombres.append(str(nombre_wb.value))


cedulas= lista["c2":"31"]
for filas in cedulas:
        for cedula_wb in filas:
                lista_cedulas.append(str(cedula_wb.value))

apellidos=lista["b2":"31"]
for filas in apellidos:
        for apellido_wb in filas:
                lista_apellidos.append(str(apellido_wb.value))

datos_alumnos=[]

a=0

while len(datos_alumnos) != len(lista_nombres):
        datos_alumnos.append((lista_nombres[a],lista_apellidos[a],lista_cedulas[a]))
                
        a+=1

print("Programa de listas - Grupo 12 JaP")


def ingresar_cedula():

        cedula=input("Ingrese su cédula: ")
        
        while (cedula not in lista_cedulas or cedula in lista_asistencia) and cedula != "0":
                print("Cédula inválida")
                cedula = input("Ingrese su cédula: ")
        if cedula =="0":
                print("Asistencias registradas")
        else:
                print("Asistencia registrada correctamente")
                
        return cedula    


def llegada_tarde():

        cedula_llegada_tarde=input("Ingrese llegada tarde: ")
        
        while (cedula_llegada_tarde  in lista_asistencia  or cedula_llegada_tarde not in lista_cedulas or cedula_llegada_tarde in lista_llegada_tarde)and cedula_llegada_tarde != "0": 
                print("Cédula inválida")
                cedula_llegada_tarde = input("Ingrese llegada tarde: ")
        
        if cedula_llegada_tarde =="0":
                print("Llegadas tarde registradas")
        else:
                print("Llegada tarde registrada correctamente")
        

        return cedula_llegada_tarde


class Alumno(object):


        def __init__(self,nombre,apellido,cedula):

                
                self.nombre=nombre
                self.apellido=apellido
                self.cedula=cedula


class Profesor(object):

        def __init__(self,nombre,apellido):

                self.nombre=nombre
                self.apellido=apellido

        def registrar_asistencia(self,grupo_alumnos):

                cedula=ingresar_cedula()
                
                while cedula!= "0":
                
                        for alumno in grupo_alumnos:
                                if cedula == alumno.cedula:
                                        lista_asistencia.append(cedula)
                                        
                        cedula=ingresar_cedula()
                        
                return lista_asistencia

        def registrar_llegada_tarde(self,grupo_alumnos):
                
                cedula_llegada_tarde=llegada_tarde()
                
                while cedula_llegada_tarde != "0":

                        for alumno in grupo_alumnos:
                                if cedula_llegada_tarde == alumno.cedula:
                                        lista_llegada_tarde.append(cedula)
                                        
                        cedula_llegada_tarde=llegada_tarde()
                        
                return lista_llegada_tarde

        
def exportar_excel(lista_asistencia,datos_alumnos,lista_llegada_tarde):

        from openpyxl import Workbook

        wb=Workbook()
        hoja=wb.active
        hoja.title="asistencias del dia"
               

        hoja["A1"]="nombre"
        hoja["B1"]="apellido"
        hoja["C1"]="cedula"
        hoja["D1"]="asistencia" 


        col=1
        filas=1


        for alumno in datos_alumnos:
                for dato in alumno:
                        hoja.cell(row=filas+1,column=col,value=dato)
                        if alumno[2] in lista_asistencia:
                                hoja.cell(row=filas+1,column=col+1,value="si")
                        elif alumno[2] in lista_llegada_tarde:
                                hoja.cell(row=filas+1,column=col+1,value="1/2")
                        elif alumno[2] not in lista_asistencia:
                                hoja.cell(row=filas+1,column=col+1,value="no")

                        col+=1
                col=1
                filas+=1
        
        
        wb.save(" asistencias del dia")




lista_asistencia=[]
lista_llegada_tarde=[]
grupo_alumnos=[]

for dat_alumno in datos_alumnos:

        alumno=Alumno(*dat_alumno)
        grupo_alumnos.append(alumno)

        

profesor=Profesor("Marco","Gentini")
profesor.registrar_asistencia(grupo_alumnos)
profesor.registrar_llegada_tarde(grupo_alumnos)
exportar_excel(lista_asistencia,datos_alumnos,lista_llegada_tarde)


        
  


                






        
  
